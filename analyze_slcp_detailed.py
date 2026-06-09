"""Detailed analysis of SLCP survey file."""

import openpyxl
import json

SLCP_SOURCE = "/Users/manju_na/Downloads/fslmsurvey_e1593509-4de9-4057-8dd7-42b639c6e297 1.xlsm"

def find_slcp_questions():
    """Find SLCP questions in the Excel file."""
    wb = openpyxl.load_workbook(SLCP_SOURCE, data_only=True)

    # Focus on likely sheets
    target_sheets = ["Main Survey Actual", "SLCP Data Collection Tool 1.7"]

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        print(f"\n{'='*60}")
        print(f"Analyzing: {sheet_name}")
        print('='*60)

        ws = wb[sheet_name]

        # Find headers by looking for common column names
        header_row = None
        for row_idx in range(1, min(20, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            row_str = ' '.join([str(v) for v in row_values if v])

            # Look for SLCP-related keywords
            if any(keyword in row_str.upper() for keyword in ['QUESTION', 'KEY', 'SECTION', 'SUBSECTION', 'NUMBER']):
                print(f"\nPotential header row {row_idx}: {row_values[:15]}")
                if not header_row:
                    header_row = row_idx

        if header_row:
            print(f"\nUsing row {header_row} as headers")
            headers = [cell.value for cell in ws[header_row]]
            print(f"Headers: {headers[:20]}")

            # Show sample data
            print(f"\nSample data rows:")
            for i in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                row = [cell.value for cell in ws[i]]
                if any(row):
                    print(f"Row {i}: {dict(zip(headers[:10], row[:10]))}")

            # Count non-empty rows
            non_empty = 0
            for i in range(header_row + 1, ws.max_row + 1):
                row = [cell.value for cell in ws[i]]
                if any(row):
                    non_empty += 1

            print(f"\nTotal non-empty data rows: {non_empty}")

        else:
            # Show first 20 rows
            print(f"\nFirst 20 rows (no clear headers found):")
            for i in range(1, min(21, ws.max_row + 1)):
                row = [cell.value for cell in ws[i]]
                if any(row):
                    print(f"Row {i}: {row[:10]}")

    wb.close()

if __name__ == "__main__":
    find_slcp_questions()
