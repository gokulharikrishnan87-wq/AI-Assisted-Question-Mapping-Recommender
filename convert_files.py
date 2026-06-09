"""Convert new RSC and SLCP files to application format."""

import json
import shutil
import openpyxl

# Source files
RSC_SOURCE = "/Users/manju_na/Downloads/RSC Questions.xlsx"
SLCP_SOURCE = "/Users/manju_na/Downloads/fslmsurvey_e1593509-4de9-4057-8dd7-42b639c6e297 1.xlsm"

# Target files
RSC_TARGET = "RSC Questions.xlsx"
SLCP_TARGET = "slcp_data_dictionary.json"

def convert_rsc():
    """Copy RSC file to project root."""
    print("Converting RSC Questions...")
    shutil.copy2(RSC_SOURCE, RSC_TARGET)

    # Verify the copy
    wb = openpyxl.load_workbook(RSC_TARGET)
    ws = wb.active
    print(f"✅ RSC file copied: {ws.max_row - 1} questions")
    wb.close()

def convert_slcp():
    """Convert SLCP Excel to JSON dictionary."""
    print("\nConverting SLCP Questions...")

    wb = openpyxl.load_workbook(SLCP_SOURCE, data_only=True)
    ws = wb["Main Survey Actual"]

    # Headers are in row 1
    headers = [cell.value for cell in ws[1]]

    # Find column indices
    key_idx = headers.index("Key")
    number_idx = headers.index("Number")
    section_idx = headers.index("Section")
    subsection_idx = headers.index("Sub-Section")
    category_idx = headers.index("Category")
    question_idx = headers.index("Main Text")
    master_type_idx = headers.index("Master Type")

    slcp_dict = {}
    question_count = 0

    # Process each row (skip header row 1, start from row 2)
    for row_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_idx]]

        key = row[key_idx]
        master_type = row[master_type_idx]
        question_text = row[question_idx]

        # Only include actual questions (not META, Instructions, Headers, etc.)
        if not key or not question_text:
            continue

        # Filter for Question types only
        if master_type not in ["Question", "CALCULATED"]:
            continue

        slcp_dict[key] = {
            "key": key,
            "number": row[number_idx] or "",
            "section": row[section_idx] or "",
            "subsection": row[subsection_idx] or "",
            "category": row[category_idx] or "",
            "question": str(question_text).strip()
        }
        question_count += 1

    wb.close()

    # Save to JSON
    with open(SLCP_TARGET, "w", encoding="utf-8") as f:
        json.dump(slcp_dict, f, indent=2, ensure_ascii=False)

    print(f"✅ SLCP file converted: {question_count} questions")
    return question_count

def verify_files():
    """Verify the converted files."""
    print("\n" + "="*60)
    print("Verification:")
    print("="*60)

    # Verify RSC
    wb = openpyxl.load_workbook(RSC_TARGET)
    ws = wb["Sheet1"]  # Use Sheet1 explicitly
    headers = [cell.value for cell in ws[1]]
    print(f"\n✅ RSC Questions.xlsx")
    print(f"   - Rows: {ws.max_row - 1}")
    print(f"   - Columns: {headers}")

    # Sample row
    sample = dict(zip(headers, [cell.value for cell in ws[2]]))
    print(f"   - Sample: {sample.get('Section', 'N/A')} / {sample.get('LLL Key (unique)', 'N/A')}")
    wb.close()

    # Verify SLCP
    with open(SLCP_TARGET) as f:
        slcp_data = json.load(f)

    print(f"\n✅ slcp_data_dictionary.json")
    print(f"   - Total questions: {len(slcp_data)}")

    # Show sample
    sample_key = list(slcp_data.keys())[0]
    sample_data = slcp_data[sample_key]
    print(f"   - Sample key: {sample_key}")
    print(f"   - Sample data: {sample_data}")

    # Count by section
    sections = {}
    for item in slcp_data.values():
        section = item.get("section", "Unknown")
        sections[section] = sections.get(section, 0) + 1

    print(f"\n   Questions by section:")
    for section, count in sorted(sections.items()):
        print(f"      - {section}: {count}")

    print("\n" + "="*60)
    print("✅ Conversion complete! Files are ready to use.")
    print("="*60)

if __name__ == "__main__":
    convert_rsc()
    convert_slcp()
    verify_files()
