"""Convert updated RSC and SLCP files to application format."""

import json
import openpyxl
import pandas as pd
from pathlib import Path

# File paths
RSC_SOURCE = "/Users/manju_na/Downloads/RSC Questions.xlsx"
SLCP_SOURCE = "/Users/manju_na/Downloads/fslmsurvey_e1593509-4de9-4057-8dd7-42b639c6e297 1.xlsm"
RSC_TARGET = "RSC Questions.xlsx"
SLCP_TARGET = "slcp_data_dictionary.json"

def analyze_rsc_file():
    """Analyze RSC Questions structure."""
    print("\n=== Analyzing RSC Questions.xlsx ===")
    wb = openpyxl.load_workbook(RSC_SOURCE)

    print(f"Sheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {ws.dimensions}")

        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers ({len(headers)}): {headers[:10]}...")

        # Get sample row
        if ws.max_row > 1:
            sample = [cell.value for cell in ws[2]]
            print(f"Sample row: {dict(zip(headers[:5], sample[:5]))}")

        print(f"Total rows: {ws.max_row}")

    wb.close()

def analyze_slcp_file():
    """Analyze SLCP survey structure."""
    print("\n=== Analyzing SLCP Survey.xlsm ===")

    try:
        # Try reading all sheets
        xl_file = pd.ExcelFile(SLCP_SOURCE)
        print(f"Sheets: {xl_file.sheet_names}")

        for sheet_name in xl_file.sheet_names[:3]:  # First 3 sheets
            print(f"\n--- Sheet: {sheet_name} ---")
            df = pd.read_excel(SLCP_SOURCE, sheet_name=sheet_name, nrows=5)
            print(f"Shape: {df.shape}")
            print(f"Columns ({len(df.columns)}): {list(df.columns)[:10]}...")
            print("\nFirst 3 rows:")
            print(df.head(3))
    except Exception as e:
        print(f"Error reading with pandas: {e}")

        # Fallback to openpyxl
        wb = openpyxl.load_workbook(SLCP_SOURCE, data_only=True)
        print(f"Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {ws.dimensions}")

            # Get headers
            headers = [cell.value for cell in ws[1]]
            print(f"Headers: {headers[:10]}...")

            # Sample rows
            for i in range(2, min(5, ws.max_row + 1)):
                row = [cell.value for cell in ws[i]]
                print(f"Row {i}: {row[:5]}...")

        wb.close()

def copy_rsc_file():
    """Copy RSC file to project root."""
    import shutil
    print(f"\n=== Copying RSC file ===")
    shutil.copy2(RSC_SOURCE, RSC_TARGET)
    print(f"✅ Copied to {RSC_TARGET}")

if __name__ == "__main__":
    analyze_rsc_file()
    analyze_slcp_file()

    # Ask before copying
    print("\n" + "="*60)
    response = input("\nCopy RSC file to project root? (y/n): ")
    if response.lower() == 'y':
        copy_rsc_file()
