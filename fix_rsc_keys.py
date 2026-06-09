"""Fix RSC LLL Keys to match description format."""

import openpyxl
import re

def fix_rsc_keys():
    """Extract proper LLL keys from descriptions."""
    wb = openpyxl.load_workbook("RSC Questions.xlsx")
    ws = wb.active

    print("Fixing LLL Keys...")
    fixed_count = 0

    # Skip header row
    for row_idx in range(2, ws.max_row + 1):
        description = ws.cell(row_idx, 3).value  # LLL Description column

        if description:
            # Extract key from description (e.g., "1.01 The facility...")
            match = re.match(r'^(\d+\.\d+)\s+', str(description))
            if match:
                proper_key = match.group(1)
                current_key = ws.cell(row_idx, 2).value  # LLL Key column

                if current_key != proper_key:
                    ws.cell(row_idx, 2).value = proper_key
                    fixed_count += 1
                    if fixed_count <= 5:
                        print(f"  Fixed row {row_idx}: '{current_key}' → '{proper_key}'")

    wb.save("RSC Questions.xlsx")
    wb.close()

    print(f"\n✅ Fixed {fixed_count} keys")

    # Verify
    wb = openpyxl.load_workbook("RSC Questions.xlsx")
    ws = wb.active

    print("\nVerification - First 5 rows:")
    for i in range(2, 7):
        key = ws.cell(i, 2).value
        desc = ws.cell(i, 3).value
        print(f"  {key}: {desc[:60]}...")

    wb.close()

if __name__ == "__main__":
    fix_rsc_keys()
