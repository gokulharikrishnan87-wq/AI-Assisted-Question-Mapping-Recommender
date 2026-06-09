#!/usr/bin/env python3
"""Load the new SLCP Excel file and check for parent questions."""

import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('/Users/manju_na/Downloads/slcp_data_dictionary (1).xlsx', data_only=True)
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")

# Get headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print(f"\nHeaders: {headers}")

# Find MS-PLA-16 (parent) and MS-PLA-16-9 (sub-option)
parent_found = False
sub_found = False

data = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not any(row):
        continue

    # Assuming structure: key, number, section, subsection, category, question
    # Adjust indices based on actual structure
    key = row[0]
    number = row[1] if len(row) > 1 else None

    if number == 'MS-PLA-16':
        print(f"\n✅ FOUND PARENT QUESTION (row {row_idx}):")
        print(f"   Key: {key}")
        print(f"   Number: {number}")
        print(f"   Row data: {row[:6]}")  # First 6 columns
        parent_found = True

    if number == 'MS-PLA-16-9':
        print(f"\n✅ FOUND SUB-OPTION (row {row_idx}):")
        print(f"   Key: {key}")
        print(f"   Number: {number}")
        print(f"   Row data: {row[:6]}")
        sub_found = True

if not parent_found:
    print("\n❌ Parent question MS-PLA-16 NOT found")
if not sub_found:
    print("\n❌ Sub-option MS-PLA-16-9 NOT found")

print(f"\n📊 Total rows (excluding header): {ws.max_row - 1}")
