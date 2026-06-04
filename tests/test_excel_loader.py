"""Tests for Excel ingestion module."""

import pytest
from pathlib import Path
import tempfile
import openpyxl

from mapper_copilot.ingestion.excel_loader import ExcelLoader
from mapper_copilot.models import RscQuestion, SlcpQuestion, ExcelColumnMap


@pytest.fixture
def temp_rsc_excel():
    """Create a temporary RSC Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header row
        ws.append(["question_id", "description", "section"])

        # Data rows
        ws.append(["1.01", "First question about licensing", "Facility Profile"])
        ws.append(["1.02", "Second question about operations", "Facility Profile"])
        ws.append(["2.01", "Third question about labor", "Labor"])

        wb.save(tmp.name)
        yield tmp.name
        Path(tmp.name).unlink()


@pytest.fixture
def temp_slcp_excel():
    """Create a temporary SLCP Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header row
        ws.append(["slcp_key", "question", "section", "subsection"])

        # Data rows
        ws.append(["fp-oc-1", "Operating license is available and up to date", "Facility Profile", "Operating Licenses"])
        ws.append(["fp-oc-2", "Business registration is valid", "Facility Profile", "Operating Licenses"])
        ws.append(["cl-1", "Child labor policy exists", "Child Labor", "Policy"])

        wb.save(tmp.name)
        yield tmp.name
        Path(tmp.name).unlink()


def test_load_rsc_questions(temp_rsc_excel):
    """Test loading RSC questions from Excel."""
    questions = ExcelLoader.load_rsc_questions(temp_rsc_excel)

    assert len(questions) == 3
    assert isinstance(questions[0], RscQuestion)
    assert questions[0].question_id == "1.01"
    assert questions[0].description == "First question about licensing"
    assert questions[0].section == "Facility Profile"


def test_load_slcp_questions(temp_slcp_excel):
    """Test loading SLCP questions from Excel."""
    questions = ExcelLoader.load_slcp_questions(temp_slcp_excel)

    assert len(questions) == 3
    assert isinstance(questions[0], SlcpQuestion)
    assert questions[0].slcp_key == "fp-oc-1"
    assert questions[0].question == "Operating license is available and up to date"
    assert questions[0].section == "Facility Profile"
    assert questions[0].subsection == "Operating Licenses"


def test_load_rsc_questions_custom_column_map(temp_rsc_excel):
    """Test loading RSC questions with custom column mapping."""
    custom_map = ExcelColumnMap(
        id_column=0,
        question_column=1,
        section_column=2,
        skip_rows=1,
    )
    questions = ExcelLoader.load_rsc_questions(temp_rsc_excel, custom_map)

    assert len(questions) == 3
    assert questions[1].question_id == "1.02"


def test_load_nonexistent_file():
    """Test that FileNotFoundError is raised for missing file."""
    with pytest.raises(FileNotFoundError):
        ExcelLoader.load_rsc_questions("/nonexistent/path/file.xlsx")


def test_empty_sheet():
    """Test loading from empty sheet (only headers)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["question_id", "description", "section"])  # Only header, no data
        wb.save(tmp.name)

        try:
            questions = ExcelLoader.load_rsc_questions(tmp.name)
            assert len(questions) == 0
        finally:
            Path(tmp.name).unlink()


def test_skip_empty_rows(temp_rsc_excel):
    """Test that empty rows are skipped gracefully."""
    # Modify the temp file to add an empty row
    wb = openpyxl.load_workbook(temp_rsc_excel)
    ws = wb.active
    ws.append([None, None, None])  # Empty row
    ws.append(["3.01", "Question after empty row", "Section"])
    wb.save(temp_rsc_excel)

    questions = ExcelLoader.load_rsc_questions(temp_rsc_excel)
    # Should have 4 questions (3 original + 1 new, skipping the empty row)
    assert len(questions) == 4
