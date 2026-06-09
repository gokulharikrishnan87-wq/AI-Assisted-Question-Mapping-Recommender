"""Update RSC Questions to 288 questions from new file."""

import openpyxl
import shutil
from pathlib import Path

SOURCE_FILE = "/Users/manju_na/Downloads/RSC Questions.xlsx"
TARGET_FILE = "RSC Questions.xlsx"
BACKUP_FILE = "RSC Questions.xlsx.backup"

def backup_current_file():
    """Backup current RSC file."""
    if Path(TARGET_FILE).exists():
        shutil.copy2(TARGET_FILE, BACKUP_FILE)
        print(f"✅ Backed up current file to {BACKUP_FILE}")

def convert_new_format():
    """Convert new RSC format to match existing structure."""
    print("\nConverting new RSC file format...")

    # Read source file
    wb_source = openpyxl.load_workbook(SOURCE_FILE)
    ws_source = wb_source["Sheet3"]

    # Create new workbook with expected structure
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "Sheet1"

    # Write headers matching original format
    # Original: ['Section', 'LLL Key (unique)', 'LLL Description', 'Reference Data', 'Severity']
    # New source: ['LLL Qu #', 'LLL Description', 'Reference Data', 'Severity']

    ws_new.append(['Section', 'LLL Key (unique)', 'LLL Description', 'Reference Data', 'Severity'])

    # Track sections and counts
    question_count = 0
    current_section = "Unknown"
    section_map = {
        1: "1. Business Ethics",
        2: "2. Child Labor",
        3: "3. Forced Labor",
        4: "4. Health & Safety",
        5: "5. Discrimination",
        6: "6. Harassment & Abuse",
        7: "7. Freedom of Association",
        8: "8. Working Hours",
        9: "9. Wages & Benefits",
        10: "10. Environment",
        11: "11. Employment Contracts",
        12: "12. Subcontracting",
        13: "13. Social Compliance",
    }

    # Process data rows (starting from row 4, after header row 3)
    for row_idx in range(4, ws_source.max_row + 1):
        lll_num = ws_source.cell(row_idx, 1).value
        lll_desc = ws_source.cell(row_idx, 2).value
        ref_data = ws_source.cell(row_idx, 3).value
        severity = ws_source.cell(row_idx, 4).value

        # Skip empty rows
        if not lll_num and not lll_desc:
            continue

        # Determine section from LLL number (e.g., 1.01 -> section 1)
        lll_key = ""
        if lll_num and isinstance(lll_num, (int, float)):
            try:
                section_num = int(str(lll_num).split('.')[0])
                current_section = section_map.get(section_num, f"{section_num}. Unknown")
                # Format LLL Key as string
                lll_key = str(lll_num).replace('.0', '')
            except (ValueError, AttributeError):
                # Keep current section if parsing fails
                lll_key = str(lll_num) if lll_num else ""

        # Write row
        ws_new.append([
            current_section,
            lll_key,
            lll_desc or "",
            ref_data or "",
            severity or ""
        ])
        question_count += 1

    # Save to target
    wb_new.save(TARGET_FILE)
    wb_source.close()
    wb_new.close()

    print(f"✅ Converted {question_count} questions to new format")
    return question_count

def verify_new_file():
    """Verify the converted file."""
    print("\n" + "="*60)
    print("Verification:")
    print("="*60)

    wb = openpyxl.load_workbook(TARGET_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"\n✅ RSC Questions.xlsx")
    print(f"   Sheet: {ws.title}")
    print(f"   Headers: {headers}")
    print(f"   Total questions: {ws.max_row - 1}")

    # Show sample rows
    print(f"\n   Sample questions:")
    for i in range(2, min(6, ws.max_row + 1)):
        row = dict(zip(headers, [cell.value for cell in ws[i]]))
        print(f"   - {row['Section']} / {row['LLL Key (unique)']}: {row['LLL Description'][:60]}...")

    # Count by section
    sections = {}
    for i in range(2, ws.max_row + 1):
        section = ws.cell(i, 1).value
        if section:
            sections[section] = sections.get(section, 0) + 1

    print(f"\n   Questions by section:")
    for section, count in sorted(sections.items()):
        print(f"      - {section}: {count}")

    wb.close()

    print("\n" + "="*60)
    print("✅ Update complete!")
    print("="*60)

if __name__ == "__main__":
    print("Updating RSC Questions to 288 questions...")
    backup_current_file()
    count = convert_new_format()
    verify_new_file()

    print(f"\n📊 Summary:")
    print(f"   Old: 829 questions")
    print(f"   New: {count} questions")
    print(f"   Backup: {BACKUP_FILE}")
